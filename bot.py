import logging
import discord
from discord.ext import commands
from discord import app_commands
from collections import OrderedDict
import json, os, random, io, datetime
from dotenv import load_dotenv
from mmr_interface import MMRHubView

from halo_bot.checks import is_admin
from halo_bot.constants import PROVISIONAL_SESSIONS, STATS_FILE, TIMEOUT_MENU, TIMEOUT_STAT
from halo_bot.pure import calculate_mmr, canonical_name, halo_rank, normalise
from halo_bot.storage import (
    MMR_FILE,
    PRESETS_FILE,
    RECALL_FILE,
    TEAM_HISTORY_FILE,
    mmr_data,
    presets,
    recall_channels,
    save_json,
    team_history,
    team_storage,
)

load_dotenv()

logging.basicConfig(
    level=getattr(logging, os.getenv("LOG_LEVEL", "INFO").upper(), logging.INFO),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
TOKEN = os.getenv("DISCORD_BOT_TOKEN")
if not TOKEN:
    raise ValueError("DISCORD_BOT_TOKEN is not set.")

intents = discord.Intents.default()
intents.members = True
intents.voice_states = True
intents.message_content = True


class HaloBot(commands.Bot):
    async def setup_hook(self):
        await self.load_extension("halo_bot.cogs.orbital")
        await self.load_extension("halo_bot.cogs.matchmaking")
        admin_cmd_names = (
            "recall",
            "teams",
            "sub",
            "import_mmr",
            "export",
            "presets",
            "history",
            "matchmaking",
            "orbital_jump",
            "sync",
        )
        for name in admin_cmd_names:
            cmd = self.tree.get_command(name)
            if cmd:
                cmd.error(_admin_error)


bot = HaloBot(command_prefix="!", intents=intents)

# None = never auto-delete (team lists, leaderboard)

def get_emoji(guild: discord.Guild, name: str) -> str:
    if guild:
        e = discord.utils.get(guild.emojis, name=name)
        if e:
            return str(e)
    return f":{name}:"

def rank_display(mmr: float, guild: discord.Guild, provisional: bool = False) -> str:
    rname, ename = halo_rank(mmr)
    remoji = get_emoji(guild, ename)
    prov = " *" if provisional else ""
    return f"{remoji} {rname}{prov}"

def is_provisional(data: dict) -> bool:
    return data.get("sessions", 0) < PROVISIONAL_SESSIONS

# ─────────────────────────────────────────────
# MMR CALCULATION (see halo_bot.pure for normalise / calculate_mmr)
# ─────────────────────────────────────────────

def parse_session_sheet(ws) -> list:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    players = []
    for row in rows[1:]:
        try:
            raw = str(row[0]).strip() if row[0] else None
            if not raw or raw.lower() in ("none", ""):
                continue
            def safe(val):
                try: return float(val or 0)
                except: return 0.0
            players.append({
                "raw_name": raw,
                "name":     canonical_name(raw),
                "kills":    safe(row[1]),
                "assists":  safe(row[2]),
                "deaths":   safe(row[3]),
                "kd":       safe(row[4]),
                "captures": safe(row[6]),
                "obj_time": safe(row[7]),
                "points":   safe(row[9]),
            })
        except:
            continue
    return players

def parse_leaderboard_sheet(ws) -> list:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    players = []
    for row in rows[1:]:
        try:
            name = str(row[1]).strip() if row[1] else None
            if not name or name.lower() == "none":
                continue
            gamertag = str(row[2]).strip() if row[2] and str(row[2]).strip().lower() != "none" else ""
            players.append({
                "name":     name,
                "gamertag": gamertag,
                "kills":    float(row[3] or 0),
                "assists":  float(row[4] or 0),
                "deaths":   float(row[5] or 0),
                "kd":       float(row[6] or 0),
                "obj_time": float(row[7] or 0),
                "captures": float(row[8] or 0),
                "points":   float(row[9] or 0),
                "sessions": int(row[10] or 1),
            })
        except:
            continue
    return players

def get_guild_mmr(guild_id: int) -> dict:
    return mmr_data.get(str(guild_id), {})

# ─────────────────────────────────────────────
# MESSAGING HELPERS
# ─────────────────────────────────────────────
class DismissView(discord.ui.View):
    def __init__(self, timeout=None):
        super().__init__(timeout=timeout)
        self.message = None

    @discord.ui.button(label="✕", style=discord.ButtonStyle.secondary)
    async def dismiss(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            await interaction.response.defer()
            await interaction.delete_original_response()
        except Exception:
            pass

    async def on_timeout(self):
        if self.message:
            try:
                await self.message.delete()
            except Exception:
                pass


async def _attach_timeout(view: DismissView, interaction: discord.Interaction, ephemeral: bool):
    """After sending, grab the message ref so on_timeout can delete it."""
    if view.timeout and not ephemeral:
        try:
            view.message = await interaction.original_response()
        except Exception:
            pass


async def send_minimal(interaction: discord.Interaction, content: str,
                       ephemeral: bool = True, timeout=None):
    view = DismissView(timeout=timeout)
    await interaction.response.send_message(content, view=view, ephemeral=ephemeral)
    await _attach_timeout(view, interaction, ephemeral)


async def followup_minimal(interaction: discord.Interaction, content: str,
                           ephemeral: bool = False, timeout=None):
    view = DismissView(timeout=timeout)
    msg  = await interaction.followup.send(content, view=view, ephemeral=ephemeral)
    if view.timeout and not ephemeral:
        view.message = msg


def chunk_lines(lines: list, header: str = "", limit: int = 1800) -> list:
    chunks, current = [], header
    for line in lines:
        candidate = (current + "\n" + line).lstrip("\n")
        if len(candidate) > limit:
            chunks.append(current)
            current = line
        else:
            current = candidate
    if current:
        chunks.append(current)
    return chunks


async def followup_chunked(interaction: discord.Interaction, lines: list,
                           header: str = "", ephemeral: bool = False, timeout=None):
    for chunk in chunk_lines(lines, header):
        view = DismissView(timeout=timeout)
        msg  = await interaction.followup.send(chunk, view=view, ephemeral=ephemeral)
        if timeout and not ephemeral:
            view.message = msg


async def send_single_or_chunked(interaction: discord.Interaction, lines: list,
                                 header: str = "", ephemeral: bool = False, timeout=None):
    message = (header + "\n" + "\n".join(lines)).strip()
    if len(message) <= 2000:
        view = DismissView(timeout=timeout)
        msg  = await interaction.followup.send(message, view=view, ephemeral=ephemeral)
        if timeout and not ephemeral:
            view.message = msg
    else:
        await followup_chunked(interaction, lines, header=header,
                               ephemeral=ephemeral, timeout=timeout)

# ─────────────────────────────────────────────
# TEAM HELPERS
# ─────────────────────────────────────────────
def find_member_team(guild_id: int, member_id: int):
    for vc_id, members in team_storage.get(guild_id, {}).items():
        if member_id in members:
            return vc_id
    return None


def build_team_summary(guild: discord.Guild, guild_id: int) -> str:
    teams = team_storage.get(guild_id, {})
    if not teams:
        return "_No teams assigned yet._"
    gmmr  = get_guild_mmr(guild_id)
    lines = ["**Current assignments**"]
    for vc_id, member_ids in teams.items():
        vc = guild.get_channel(int(vc_id))
        names, mmr_vals = [], []
        for mid in member_ids:
            m     = guild.get_member(mid)
            dname = m.display_name if m else f"Unknown({mid})"
            cname = canonical_name(dname)
            pdata = gmmr.get(cname) or gmmr.get(dname)
            if pdata:
                rd = rank_display(pdata["mmr"], guild, is_provisional(pdata))
                names.append(f"{dname} {rd}({pdata['mmr']})")
                mmr_vals.append(pdata["mmr"])
            else:
                names.append(dname)
        avg  = f" · avg **{round(sum(mmr_vals)/len(mmr_vals), 1)}** MMR" if mmr_vals else ""
        lines.append(f"**{vc.name if vc else vc_id}** · {len(names)} players{avg}\n└ {', '.join(names)}")
    return "\n".join(lines)


def team_builder_member_buckets(guild: discord.Guild):
    """Split non-bot members: in voice, online (not in voice), offline/invisible."""
    in_voice = {}
    for vc in guild.voice_channels:
        for m in vc.members:
            if not m.bot:
                in_voice[m.id] = m
    in_ids = set(in_voice.keys())
    online_nv, offline = [], []
    for m in guild.members:
        if m.bot or m.id in in_ids:
            continue
        if m.status in (discord.Status.online, discord.Status.idle, discord.Status.dnd):
            online_nv.append(m)
        else:
            offline.append(m)
    sk = lambda m: m.display_name.lower()
    voice_sorted = sorted(in_voice.values(), key=sk)
    online_sorted = sorted(online_nv, key=sk)
    off_sorted = sorted(offline, key=sk)
    return (
        voice_sorted[:25],
        online_sorted[:25],
        off_sorted[:25],
        len(voice_sorted),
        len(online_sorted),
        len(off_sorted),
    )


def _select_options_from_members(members: list, desc_prefix: str) -> list:
    opts = []
    for m in members:
        label = m.display_name[:80]
        if len(m.display_name) > 80:
            label = m.display_name[:77] + "…"
        opts.append(
            discord.SelectOption(
                label=label,
                value=str(m.id),
                description=f"{desc_prefix} · {m.status.name}",
            )
        )
    return opts


def save_team_to_history(guild_id: int, guild: discord.Guild, label: str = None):
    gid  = str(guild_id)
    if gid not in team_history:
        team_history[gid] = []
    teams = team_storage.get(guild_id, {})
    if not teams:
        return
    snapshot = {}
    for vc_id, member_ids in teams.items():
        vc      = guild.get_channel(int(vc_id))
        vc_name = vc.name if vc else str(vc_id)
        names   = [guild.get_member(mid).display_name
                   if guild.get_member(mid) else str(mid) for mid in member_ids]
        snapshot[vc_name] = names
    entry = {
        "label": label or datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "teams": snapshot,
    }
    team_history[gid].insert(0, entry)
    team_history[gid] = team_history[gid][:10]
    save_json(TEAM_HISTORY_FILE, team_history)

# ─────────────────────────────────────────────
# RANDOMISE VIEW
# ─────────────────────────────────────────────
class RandomiseView(discord.ui.View):
    def __init__(self, guild: discord.Guild):
        super().__init__(timeout=TIMEOUT_MENU)
        self.guild = guild
        vcs = [c for c in guild.channels if isinstance(c, discord.VoiceChannel)]
        self.channel_select = RandomChannelSelect(vcs)
        self.add_item(self.channel_select)

    @discord.ui.button(label="🎲 Randomise!", style=discord.ButtonStyle.success, row=1)
    async def randomise(self, interaction: discord.Interaction, button: discord.ui.Button):
        selected = self.channel_select.selected_ids
        if len(selected) < 2:
            await interaction.response.edit_message(
                content="🎲 **Randomise** — Pick at least 2 channels first.", view=self)
            return
        all_members = list({m.id: m for vc in self.guild.voice_channels
                            for m in vc.members}.values())
        if not all_members:
            await interaction.response.edit_message(
                content="🎲 **Randomise** — No members in voice.", view=self)
            return
        random.shuffle(all_members)
        buckets = {vc_id: [] for vc_id in selected}
        for i, m in enumerate(all_members):
            buckets[selected[i % len(selected)]].append(m)
        gid = self.guild.id
        team_storage[gid] = {}
        await interaction.response.defer()
        results = []
        for vc_id, members in buckets.items():
            vc = self.guild.get_channel(int(vc_id))
            if not vc:
                continue
            team_storage[gid][vc_id] = []
            moved, skipped = [], []
            for m in members:
                team_storage[gid][vc_id].append(m.id)
                if m.voice:
                    await m.move_to(vc)
                    moved.append(m.display_name)
                else:
                    skipped.append(m.display_name)
            line = f"**{vc.name}** ({len(moved)}): {', '.join(moved) or 'nobody'}"
            if skipped:
                line += f" _(not in voice: {', '.join(skipped)})_"
            results.append(line)
        save_team_to_history(gid, self.guild, "🎲 Random")
        await interaction.followup.send(
            "🎲 **Teams randomised!**\n" + "\n".join(results),
            view=DismissView(), ephemeral=False)

    @discord.ui.button(label="✕ Cancel", style=discord.ButtonStyle.secondary, row=1)
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(content="Cancelled.", view=None)


class RandomChannelSelect(discord.ui.Select):
    def __init__(self, vcs):
        self.selected_ids: list = []
        options = [discord.SelectOption(label=c.name,
                   description=f"{len(c.members)} connected", value=str(c.id))
                   for c in vcs[:25]]
        super().__init__(placeholder="Pick 2–25 voice channels...",
                         options=options, min_values=2, max_values=min(len(vcs), 25), row=0)

    async def callback(self, interaction: discord.Interaction):
        self.selected_ids = self.values
        names = [interaction.guild.get_channel(int(v)).name
                 for v in self.values if interaction.guild.get_channel(int(v))]
        await interaction.response.edit_message(
            content=f"🎲 **Randomise** — Teams: **{', '.join(names)}** — click 🎲 Randomise!",
            view=self.view)

# ─────────────────────────────────────────────
# BALANCED MATCHMAKING VIEW
# ─────────────────────────────────────────────
class MatchmakeView(discord.ui.View):
    def __init__(self, guild: discord.Guild):
        super().__init__(timeout=TIMEOUT_MENU)
        self.guild = guild
        vcs = [c for c in guild.channels if isinstance(c, discord.VoiceChannel)]
        self.channel_select = MatchmakeChannelSelect(vcs)
        self.add_item(self.channel_select)

    @discord.ui.button(label="⚖️ Generate Balanced Teams", style=discord.ButtonStyle.success, row=1)
    async def matchmake(self, interaction: discord.Interaction, button: discord.ui.Button):
        selected = self.channel_select.selected_ids
        if len(selected) < 2:
            await interaction.response.edit_message(
                content="⚖️ **Balanced** — Pick at least 2 channels first.", view=self)
            return
        all_members = list({m.id: m for vc in self.guild.voice_channels
                            for m in vc.members}.values())
        if not all_members:
            await interaction.response.edit_message(
                content="⚖️ **Balanced** — No members in voice.", view=self)
            return
        gmmr = get_guild_mmr(self.guild.id)
        num_teams = len(selected)
        rated, unrated = [], []
        for m in all_members:
            cname = canonical_name(m.display_name)
            pdata = gmmr.get(cname) or gmmr.get(m.display_name)
            if pdata:
                rated.append((m, pdata["mmr"], pdata))
            else:
                unrated.append(m)
        rated.sort(key=lambda x: x[1], reverse=True)
        buckets   = {vc_id: [] for vc_id in selected}
        direction, idx = 1, 0
        for member, mmr, pdata in rated:
            buckets[selected[idx]].append((member, mmr, pdata))
            idx += direction
            if idx >= num_teams:
                idx = num_teams - 1; direction = -1
            elif idx < 0:
                idx = 0; direction = 1
        random.shuffle(unrated)
        for i, m in enumerate(unrated):
            buckets[selected[i % num_teams]].append((m, None, None))
        gid = self.guild.id
        team_storage[gid] = {}
        await interaction.response.defer()
        results = []
        for vc_id, members in buckets.items():
            vc = self.guild.get_channel(int(vc_id))
            if not vc:
                continue
            team_storage[gid][vc_id] = []
            moved, skipped, mmr_vals = [], [], []
            for m, mmr, pdata in members:
                team_storage[gid][vc_id].append(m.id)
                if mmr is not None:
                    rd = rank_display(mmr, self.guild, is_provisional(pdata))
                    label = f"{m.display_name} {rd}"
                    mmr_vals.append(mmr)
                else:
                    label = f"{m.display_name} ❔"
                if m.voice:
                    await m.move_to(vc)
                    moved.append(label)
                else:
                    skipped.append(label)
            avg  = f" | avg MMR: {round(sum(mmr_vals)/len(mmr_vals),1)}" if mmr_vals else ""
            line = f"**{vc.name}**{avg}: {', '.join(moved) or 'nobody'}"
            if skipped:
                line += f" _(not in voice: {', '.join(skipped)})_"
            results.append(line)
        save_team_to_history(gid, self.guild, "⚖️ Balanced")
        await interaction.followup.send(
            "⚖️ **Balanced teams generated!**\n" + "\n".join(results),
            view=DismissView(), ephemeral=False)

    @discord.ui.button(label="✕ Cancel", style=discord.ButtonStyle.secondary, row=1)
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(content="Cancelled.", view=None)


class MatchmakeChannelSelect(discord.ui.Select):
    def __init__(self, vcs):
        self.selected_ids: list = []
        options = [discord.SelectOption(label=c.name,
                   description=f"{len(c.members)} connected", value=str(c.id))
                   for c in vcs[:25]]
        super().__init__(placeholder="Pick 2–25 voice channels...",
                         options=options, min_values=2, max_values=min(len(vcs), 25), row=0)

    async def callback(self, interaction: discord.Interaction):
        self.selected_ids = self.values
        names = [interaction.guild.get_channel(int(v)).name
                 for v in self.values if interaction.guild.get_channel(int(v))]
        await interaction.response.edit_message(
            content=f"⚖️ **Balanced** — Channels: **{', '.join(names)}** — click ⚖️ to generate!",
            view=self.view)

# ─────────────────────────────────────────────
# PRESET VIEWS
# ─────────────────────────────────────────────
class SavePresetModal(discord.ui.Modal, title="Save Team Preset"):
    preset_name = discord.ui.TextInput(label="Preset Name",
                  placeholder="e.g. Monday Night 6v6", max_length=50)
    preset_note = discord.ui.TextInput(label="Notes (optional)",
                  placeholder="e.g. Competitive lineup", required=False,
                  max_length=150, style=discord.TextStyle.paragraph)

    def __init__(self, guild):
        super().__init__()
        self.guild = guild

    async def on_submit(self, interaction: discord.Interaction):
        gid   = str(self.guild.id)
        teams = team_storage.get(self.guild.id, {})
        if not teams:
            await interaction.response.send_message(
                "⚠️ No active teams to save.", ephemeral=True)
            return
        if gid not in presets:
            presets[gid] = {}
        snapshot = {}
        for vc_id, member_ids in teams.items():
            vc      = self.guild.get_channel(int(vc_id))
            vc_name = vc.name if vc else str(vc_id)
            names   = [self.guild.get_member(mid).display_name
                       if self.guild.get_member(mid) else str(mid) for mid in member_ids]
            snapshot[vc_name] = names
        name = str(self.preset_name)
        presets[gid][name] = {
            "note": str(self.preset_note) if self.preset_note.value else "",
            "teams": snapshot,
        }
        save_json(PRESETS_FILE, presets)
        await interaction.response.send_message(
            f"✅ Preset **{name}** saved!", ephemeral=True)


class TeamPresetsView(discord.ui.View):
    def __init__(self, guild):
        super().__init__(timeout=TIMEOUT_MENU)
        self.guild = guild
        gp = presets.get(str(guild.id), {})
        if gp:
            self.preset_select = TeamPresetSelect(gp)
            self.add_item(self.preset_select)

    @discord.ui.button(label="📂 Presets", style=discord.ButtonStyle.success, row=1)
    async def load_preset(self, interaction: discord.Interaction, button: discord.ui.Button):
        preset_name = self.preset_select.selected_name
        if not preset_name:
            await interaction.response.edit_message(
                content="📂 **Load Preset** — Select a preset first.", view=self)
            return
        preset = presets.get(str(self.guild.id), {}).get(preset_name)
        if not preset:
            await interaction.response.edit_message(
                content=f"⚠️ Preset **{preset_name}** not found.", view=self)
            return
        vcs_by_name     = {vc.name: vc for vc in self.guild.voice_channels}
        members_by_name = {m.display_name: m for m in self.guild.members}
        gid = self.guild.id
        team_storage[gid] = {}
        loaded, missing_ch, missing_m = [], [], []
        for vc_name, member_names in preset["teams"].items():
            vc = vcs_by_name.get(vc_name)
            if not vc:
                missing_ch.append(vc_name)
                continue
            team_storage[gid][str(vc.id)] = []
            found = []
            for mname in member_names:
                m = members_by_name.get(mname)
                if m:
                    team_storage[gid][str(vc.id)].append(m.id)
                    found.append(mname)
                else:
                    missing_m.append(mname)
            loaded.append(f"**{vc_name}**: {', '.join(found) or 'nobody'}")
        msg = f"📂 **{preset_name}** loaded!\n" + "\n".join(loaded)
        if missing_ch:
            msg += f"\n⚠️ Channels not found: {', '.join(missing_ch)}"
        if missing_m:
            msg += f"\n⚠️ Members not found: {', '.join(missing_m)}"
        msg += "\n\nUse 🚀 Send Teams to move everyone."
        await interaction.response.edit_message(content=msg, view=self)


class TeamPresetSelect(discord.ui.Select):
    def __init__(self, gp):
        self.selected_name = None
        options = [discord.SelectOption(label=n,
                   description=d.get("note", "")[:50] or "No notes", value=n)
                   for n, d in list(gp.items())[:25]]
        super().__init__(placeholder="Choose a preset...", options=options, row=0)

    async def callback(self, interaction: discord.Interaction):
        self.selected_name = self.values[0]
        preset = presets.get(str(interaction.guild_id), {}).get(self.selected_name, {})
        lines  = [f"📋 **{self.selected_name}**"]
        if preset.get("note"):
            lines.append(f"_{preset['note']}_")
        for vc_name, members in preset.get("teams", {}).items():
            lines.append(f"**{vc_name}**: {', '.join(members)}")
        lines.append("\nClick 📂 Load Preset to activate.")
        await interaction.response.edit_message(content="\n".join(lines), view=self.view)

# ─────────────────────────────────────────────
# TEAM HISTORY VIEW
# ─────────────────────────────────────────────
class TeamHistoryView(discord.ui.View):
    def __init__(self, guild):
        super().__init__(timeout=TIMEOUT_MENU)
        history = team_history.get(str(guild.id), [])
        if history:
            self.add_item(TeamHistorySelect(history))


class TeamHistorySelect(discord.ui.Select):
    def __init__(self, history):
        options = [discord.SelectOption(label=e["label"][:50], value=str(i))
                   for i, e in enumerate(history[:25])]
        super().__init__(placeholder="Choose a past configuration...", options=options, row=0)

    async def callback(self, interaction: discord.Interaction):
        history = team_history.get(str(interaction.guild_id), [])
        idx     = int(self.values[0])
        if idx >= len(history):
            await interaction.response.edit_message(
                content="⚠️ Entry not found.", view=self.view)
            return
        entry = history[idx]
        lines = [f"📜 **{entry['label']}**\n"]
        for vc_name, members in entry["teams"].items():
            lines.append(f"**{vc_name}** ({len(members)}): {', '.join(members)}")
        await interaction.response.edit_message(
            content="\n".join(lines), view=self.view)

# ─────────────────────────────────────────────
# RECALL PICKER (used inside Team Builder)
# ─────────────────────────────────────────────
class RecallPickerView(discord.ui.View):
    """Used from team builder 🔁 button.
    If a lobby is saved: shows a Confirm button.
    Otherwise: shows a channel picker select."""
    def __init__(self, guild):
        super().__init__(timeout=TIMEOUT_MENU)
        self.guild = guild
        saved_id = recall_channels.get(str(guild.id))
        self.saved_lobby = guild.get_channel(saved_id) if saved_id else None

        if self.saved_lobby:
            # Add a confirm button — clean single click to recall
            btn = discord.ui.Button(
                label=f"✅ Confirm — recall to {self.saved_lobby.name}",
                style=discord.ButtonStyle.success, row=0)
            btn.callback = self.on_confirm
            self.add_item(btn)
            # Add a change button
            change_btn = discord.ui.Button(
                label="🔄 Change lobby", style=discord.ButtonStyle.secondary, row=0)
            change_btn.callback = self.on_change
            self.add_item(change_btn)
        else:
            options = [discord.SelectOption(label=c.name, value=str(c.id))
                       for c in guild.voice_channels[:25]]
            self.select = discord.ui.Select(
                placeholder="Pick the lobby channel...", options=options, row=0)
            self.select.callback = self.on_select
            self.add_item(self.select)

    async def _do_recall(self, interaction: discord.Interaction, lobby: discord.VoiceChannel):
        recall_channels[str(self.guild.id)] = lobby.id
        save_json(RECALL_FILE, recall_channels)
        moved = 0
        for vc in self.guild.voice_channels:
            if vc.id == lobby.id:
                continue
            for member in list(vc.members):
                if member.voice:
                    await member.move_to(lobby)
                    moved += 1
        await interaction.response.edit_message(
            content=f"✅ Recalled **{moved}** member(s) to **{lobby.name}**.", view=None)

    async def on_confirm(self, interaction: discord.Interaction):
        await self._do_recall(interaction, self.saved_lobby)

    async def on_change(self, interaction: discord.Interaction):
        options = [discord.SelectOption(label=c.name, value=str(c.id))
                   for c in self.guild.voice_channels[:25]]
        self.clear_items()
        self.select = discord.ui.Select(
            placeholder="Pick a new lobby channel...", options=options, row=0)
        self.select.callback = self.on_select
        self.add_item(self.select)
        await interaction.response.edit_message(
            content="🔁 **Recall** — pick a new lobby channel:", view=self)

    async def on_select(self, interaction: discord.Interaction):
        lobby = self.guild.get_channel(int(self.select.values[0]))
        if not lobby:
            await interaction.response.edit_message(content="⚠️ Channel not found.", view=self)
            return
        await self._do_recall(interaction, lobby)

# ─────────────────────────────────────────────
# TEAM BUILDER
# ─────────────────────────────────────────────
class TeamChannelSelect(discord.ui.Select):
    def __init__(self, vcs, row: int = 0):
        self.selected_vc_id = None
        options = [discord.SelectOption(label=c.name, value=str(c.id)) for c in vcs[:25]]
        super().__init__(
            placeholder="1️⃣ Team voice channel…",
            options=options or [discord.SelectOption(label="No channels", value="none")],
            min_values=1,
            max_values=1,
            row=row,
        )

    async def callback(self, interaction: discord.Interaction):
        if self.values == ["none"]:
            await interaction.response.defer()
            return
        self.selected_vc_id = self.values[0]
        vc = interaction.guild.get_channel(int(self.selected_vc_id))
        builder = self.view
        await interaction.response.edit_message(
            content=builder._builder_content(
                f"✅ Target channel: **{vc.name if vc else '?'}** — add people, then **➕ Assign**."
            ),
            view=builder,
        )


class StagingUserSelect(discord.ui.UserSelect):
    """Native Discord multiselect — add up to 25 members per pick (accumulates)."""

    def __init__(self, row: int = 1):
        super().__init__(
            placeholder="2️⃣ Add members (picker, up to 25 at once)…",
            min_values=1,
            max_values=25,
            row=row,
        )

    async def callback(self, interaction: discord.Interaction):
        builder: TeamBuilderView = self.view
        n = 0
        for m in self.values:
            if isinstance(m, discord.Member) and not m.bot:
                builder.staged_member_ids.add(m.id)
                n += 1
        await interaction.response.edit_message(
            content=builder._builder_content(
                f"✅ Added **{n}** from picker · **{len(builder.staged_member_ids)}** total staged."
            ),
            view=builder,
        )


class MemberCategoryPickView(discord.ui.View):
    """Multiselect from in-voice / online / offline buckets (max 25 each list)."""

    def __init__(self, builder: "TeamBuilderView"):
        super().__init__(timeout=TIMEOUT_MENU)
        self.builder = builder
        self.guild = builder.guild
        voice, online, off, tv, to, tf = team_builder_member_buckets(self.guild)
        self._totals = (tv, to, tf)
        row = 0
        if voice:
            self.add_item(self._make_cat_select("🎙️ In voice", voice, "voice", row))
            row += 1
        if online:
            self.add_item(self._make_cat_select(
                "🟢 Online (not in voice)", online, "online", row))
            row += 1
        if off:
            self.add_item(self._make_cat_select("⚫ Offline / invisible", off, "away", row))
            row += 1

    def _make_cat_select(self, placeholder, members, desc_pfx, row):
        opts = _select_options_from_members(members, desc_pfx)
        mx = min(len(opts), 25)

        class CatSel(discord.ui.Select):
            def __init__(inner_self):
                super().__init__(
                    placeholder=placeholder[:150],
                    options=opts,
                    min_values=0,
                    max_values=mx,
                    row=row,
                )

            async def callback(inner_self, interaction: discord.Interaction):
                await interaction.response.defer()

        return CatSel()

    def _header(self) -> str:
        tv, to, tf = self._totals
        parts = [
            "### 📂 Pick from lists",
            "Each menu is **multi-select** (0–25 per list). Then tap **✅ Add to staging**.",
            f"· 🎙️ In voice — **up to 25** shown of **{tv}** total",
            f"· 🟢 Online, not in voice — **up to 25** of **{to}**",
            f"· ⚫ Offline / invisible — **up to 25** of **{tf}**",
        ]
        if tv == 0 and to == 0 and tf == 0:
            parts.append("\n_No non-bot members found._")
        return "\n".join(parts)

    @discord.ui.button(label="✅ Add to staging", style=discord.ButtonStyle.success, row=3)
    async def apply_picks(self, interaction: discord.Interaction, button: discord.ui.Button):
        merged = set()
        for child in self.children:
            if isinstance(child, discord.ui.Select):
                merged |= {int(x) for x in child.values}
        if not merged:
            await interaction.response.send_message(
                "⚠️ Select at least one member in the lists, or use **✕ Close**.",
                ephemeral=True,
            )
            return
        before = len(self.builder.staged_member_ids)
        self.builder.staged_member_ids.update(merged)
        added = len(self.builder.staged_member_ids) - before
        await self.builder.sync_main_panel()
        await interaction.response.edit_message(
            content=f"✅ Staged **{added}** more (**{len(self.builder.staged_member_ids)}** total on main panel).",
            view=None,
        )

    @discord.ui.button(label="✕ Close", style=discord.ButtonStyle.secondary, row=3)
    async def close_pick(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(content="Closed.", view=None)


class TeamBuilderToolsView(discord.ui.View):
    """History / recall / clear-all without crowding the main panel."""

    def __init__(self, builder: "TeamBuilderView"):
        super().__init__(timeout=TIMEOUT_MENU)
        self.builder = builder

    @discord.ui.button(label="🗑️ Clear all team slots", style=discord.ButtonStyle.danger, row=0)
    async def clear_all_teams(self, interaction: discord.Interaction, button: discord.ui.Button):
        team_storage.pop(self.builder.guild.id, None)
        await self.builder.sync_main_panel(self.builder._builder_content("✅ All team slots cleared."))
        await interaction.response.edit_message(
            content="Cleared every team on the main panel.", view=None)

    @discord.ui.button(label="📜 Team history", style=discord.ButtonStyle.primary, row=0)
    async def open_history(self, interaction: discord.Interaction, button: discord.ui.Button):
        gid = str(self.builder.guild.id)
        if not team_history.get(gid):
            await interaction.response.send_message("⚠️ No team history yet.", ephemeral=True)
            return
        await interaction.response.edit_message(content="Sent history in a follow-up ⬇️", view=None)
        await interaction.followup.send(
            "📜 **Team History** — select an entry:",
            view=TeamHistoryView(self.builder.guild),
            ephemeral=True,
        )

    @discord.ui.button(label="🔁 Recall to lobby", style=discord.ButtonStyle.secondary, row=0)
    async def open_recall(self, interaction: discord.Interaction, button: discord.ui.Button):
        saved_id = recall_channels.get(str(self.builder.guild.id))
        saved = self.builder.guild.get_channel(saved_id) if saved_id else None
        msg = (
            f"🔁 **Recall** — lobby is **{saved.name}**"
            if saved
            else "🔁 **Recall** — pick a lobby channel:"
        )
        await interaction.response.edit_message(content="Sent recall picker ⬇️", view=None)
        await interaction.followup.send(msg, view=RecallPickerView(self.builder.guild), ephemeral=True)


class TeamBuilderView(discord.ui.View):
    def __init__(self, guild: discord.Guild):
        super().__init__(timeout=None)
        self.guild = guild
        self.message: discord.WebhookMessage | discord.Message | None = None
        self.staged_member_ids: set = set()
        vcs = [c for c in guild.channels if isinstance(c, discord.VoiceChannel)]
        self.channel_select = TeamChannelSelect(vcs, row=0)
        self.user_select = StagingUserSelect(row=1)
        self.add_item(self.channel_select)
        self.add_item(self.user_select)

    async def sync_main_panel(self, content: str = None):
        text = content if content is not None else self._builder_content()
        if self.message:
            await self.message.edit(content=text, view=self)

    def _staged_line(self) -> str:
        if not self.staged_member_ids:
            return "_**Staged:** none — use the member picker or **📂 From lists**._"
        names = []
        for mid in sorted(self.staged_member_ids):
            m = self.guild.get_member(mid)
            names.append(m.display_name if m else str(mid))
        preview = ", ".join(names[:10])
        if len(names) > 10:
            preview += f" … *+{len(names) - 10} more*"
        return f"**Staged ({len(names)}):** {preview}"

    def _builder_content(self, status: str = "") -> str:
        saved_id = recall_channels.get(str(self.guild.id))
        saved = self.guild.get_channel(saved_id) if saved_id else None
        lobby = f"**{saved.name}**" if saved else "_not set — `/recall`_"
        ch = self.channel_select.selected_vc_id
        ch_name = "_pick above ↑_"
        if ch:
            vc = self.guild.get_channel(int(ch))
            ch_name = f"**{vc.name}**" if vc else str(ch)
        base = (
            "## 👥 Team Builder\n"
            f"{self._staged_line()}\n"
            f"**Target channel:** {ch_name}\n"
            "—\n"
            "**Flow:** (1) voice channel · (2) add people · (3) **➕ Assign**\n"
            "· **📂 From lists** — multiselect 🎙️ voice / 🟢 online / ⚫ offline\n"
            "· **🔊 +All in voice** — stage everyone currently in a voice channel\n"
            "**Auto:** 🎲 Random · ⚖️ MMR balance · **💾** Save · **📂** Presets\n"
            f"**Lobby:** {lobby}\n"
            "—\n"
            "**⚙️ More** — clear all slots, history, recall"
        )
        if status:
            base += f"\n\n{status}"
        return base

    @discord.ui.button(label="📂 From lists", style=discord.ButtonStyle.primary, row=2)
    async def open_lists(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = MemberCategoryPickView(self)
        await interaction.response.send_message(view._header(), view=view, ephemeral=True)

    @discord.ui.button(label="🧹 Clear staged", style=discord.ButtonStyle.secondary, row=2)
    async def clear_staged(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.staged_member_ids.clear()
        await interaction.response.edit_message(
            content=self._builder_content("✅ Staging cleared."), view=self)

    @discord.ui.button(label="🔊 +All in voice", style=discord.ButtonStyle.secondary, row=2)
    async def add_all_voice(self, interaction: discord.Interaction, button: discord.ui.Button):
        seen = {}
        for vc in self.guild.voice_channels:
            for m in vc.members:
                if not m.bot:
                    seen[m.id] = m
        for mid in seen:
            self.staged_member_ids.add(mid)
        await interaction.response.edit_message(
            content=self._builder_content(
                f"✅ Staged **{len(seen)}** from voice ({len(self.staged_member_ids)} unique total)."
            ),
            view=self,
        )

    @discord.ui.button(label="➕ Assign", style=discord.ButtonStyle.success, row=3)
    async def assign(self, interaction: discord.Interaction, button: discord.ui.Button):
        gid = self.guild.id
        vc_id = self.channel_select.selected_vc_id
        if not self.staged_member_ids:
            await interaction.response.edit_message(
                content=self._builder_content("⚠️ Stage members first (picker, lists, or **+All in voice**)."),
                view=self,
            )
            return
        if not vc_id:
            await interaction.response.edit_message(
                content=self._builder_content("⚠️ Pick a **team voice channel** first (top menu)."),
                view=self,
            )
            return
        if gid not in team_storage:
            team_storage[gid] = {}
        if vc_id not in team_storage[gid]:
            team_storage[gid][vc_id] = []
        added, moved_from = [], []
        for mid in list(self.staged_member_ids):
            m = self.guild.get_member(mid)
            if not m:
                continue
            prev = find_member_team(gid, m.id)
            if prev and prev != vc_id:
                team_storage[gid][prev].remove(m.id)
                prev_vc = self.guild.get_channel(int(prev))
                moved_from.append(
                    f"{m.display_name} (was {prev_vc.name if prev_vc else prev})")
            if m.id not in team_storage[gid][vc_id]:
                team_storage[gid][vc_id].append(m.id)
                added.append(m.display_name)
        self.staged_member_ids.clear()
        vc = self.guild.get_channel(int(vc_id))
        status = f"✅ **{vc.name if vc else vc_id}** ← {', '.join(added) or '—'}"
        if moved_from:
            status += f"\n🔄 Reassigned: {', '.join(moved_from)}"
        status += f"\n\n{build_team_summary(self.guild, gid)}"
        await interaction.response.edit_message(content=self._builder_content(status), view=self)

    @discord.ui.button(label="🚀 Send", style=discord.ButtonStyle.success, row=3)
    async def send_teams(self, interaction: discord.Interaction, button: discord.ui.Button):
        gid   = self.guild.id
        teams = team_storage.get(gid, {})
        if not teams:
            await interaction.response.edit_message(
                content=self._builder_content("⚠️ No teams assigned yet."), view=self)
            return
        await interaction.response.defer()
        results = []
        for vc_id, member_ids in teams.items():
            vc = self.guild.get_channel(int(vc_id))
            if not vc:
                continue
            moved, skipped = [], []
            for mid in member_ids:
                m = self.guild.get_member(mid)
                if m and m.voice:
                    await m.move_to(vc)
                    moved.append(m.display_name)
                elif m:
                    skipped.append(m.display_name)
            line = f"**{vc.name}**: {', '.join(moved) or 'nobody moved'}"
            if skipped:
                line += f" _(not in voice: {', '.join(skipped)})_"
            results.append(line)
        save_team_to_history(gid, self.guild)
        await interaction.followup.send(
            "## 🚀 Teams dispatched\n" + "\n".join(results),
            view=DismissView(), ephemeral=False)

    @discord.ui.button(label="📋 Teams", style=discord.ButtonStyle.secondary, row=3)
    async def show_teams(self, interaction: discord.Interaction, button: discord.ui.Button):
        summary = build_team_summary(self.guild, self.guild.id)
        await interaction.response.edit_message(
            content=self._builder_content(summary), view=self)

    @discord.ui.button(label="⚙️ More", style=discord.ButtonStyle.secondary, row=3)
    async def more_tools(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            "### ⚙️ Team Builder — more",
            view=TeamBuilderToolsView(self),
            ephemeral=True,
        )

    @discord.ui.button(label="🎲 Randomise", style=discord.ButtonStyle.primary, row=4)
    async def randomise_teams(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            "### 🎲 Randomise teams\nPick **2–25** voice channels, then confirm.",
            view=RandomiseView(self.guild), ephemeral=True)

    @discord.ui.button(label="⚖️ Balanced", style=discord.ButtonStyle.primary, row=4)
    async def balanced_teams(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            "### ⚖️ MMR-balanced teams\nPick channels to fill.",
            view=MatchmakeView(self.guild), ephemeral=True)

    @discord.ui.button(label="💾 Save", style=discord.ButtonStyle.secondary, row=4)
    async def save_preset(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(SavePresetModal(self.guild))

    @discord.ui.button(label="📂 Presets", style=discord.ButtonStyle.secondary, row=4)
    async def load_preset(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not presets.get(str(self.guild.id)):
            await interaction.response.send_message(
                "⚠️ No presets saved yet.", ephemeral=True)
            return
        await interaction.response.send_message(
            "### 📂 Load preset\nSelect a saved lineup:",
            view=TeamPresetsView(self.guild), ephemeral=True)

# ─────────────────────────────────────────────
# SLASH COMMANDS
# ─────────────────────────────────────────────

@bot.tree.command(name="recall",
    description="[Admin] Set the lobby channel and recall all members to it.")
@is_admin()
async def recall(interaction: discord.Interaction, lobby: discord.VoiceChannel):
    gid = str(interaction.guild_id)
    recall_channels[gid] = lobby.id
    save_json(RECALL_FILE, recall_channels)
    await interaction.response.defer()
    moved = 0
    for vc in interaction.guild.voice_channels:
        if vc.id == lobby.id:
            continue
        for member in list(vc.members):
            if member.voice:
                await member.move_to(lobby)
                moved += 1
    await interaction.followup.send(
        f"✅ **{lobby.name}** set as lobby. Recalled **{moved}** member(s).",
        view=DismissView(), ephemeral=False)


@bot.tree.command(name="teams",
    description="[Admin] Build teams and send members to voice channels.")
@is_admin()
async def teams(interaction: discord.Interaction):
    if not [c for c in interaction.guild.channels if isinstance(c, discord.VoiceChannel)]:
        await interaction.response.send_message(
            "⚠️ No voice channels found.", ephemeral=True)
        return
    view = TeamBuilderView(interaction.guild)
    await interaction.response.send_message(
        view._builder_content(), view=view, ephemeral=False)
    view.message = await interaction.original_response()


@bot.tree.command(name="sub",
    description="[Admin] Swap two players between teams.")
@is_admin()
async def sub(interaction: discord.Interaction, player_out: str, player_in: str):
    gid   = interaction.guild.id
    teams = team_storage.get(gid, {})
    if not teams:
        await interaction.response.send_message(
            "⚠️ No active teams. Use `/teams` first.", ephemeral=True)
        return
    guild = interaction.guild
    member_out = next(
        (guild.get_member(mid) for mids in teams.values() for mid in mids
         if guild.get_member(mid)
         and guild.get_member(mid).display_name.lower() == player_out.lower()), None)
    if not member_out:
        await interaction.response.send_message(
            f"⚠️ **{player_out}** not found in any team.", ephemeral=True)
        return
    member_in = next(
        (m for m in guild.members if m.display_name.lower() == player_in.lower()), None)
    if not member_in:
        await interaction.response.send_message(
            f"⚠️ **{player_in}** not found in this server.", ephemeral=True)
        return
    out_vc_id = find_member_team(gid, member_out.id)
    if not out_vc_id:
        await interaction.response.send_message(
            f"⚠️ **{player_out}** is not assigned to a team.", ephemeral=True)
        return
    team_storage[gid][out_vc_id].remove(member_out.id)
    in_vc_id = find_member_team(gid, member_in.id)
    if in_vc_id:
        team_storage[gid][in_vc_id].remove(member_in.id)
    team_storage[gid][out_vc_id].append(member_in.id)
    vc         = guild.get_channel(int(out_vc_id))
    voice_note = ""
    if member_in.voice:
        await member_in.move_to(vc)
        voice_note = " and moved to voice"
    await interaction.response.send_message(
        f"## 🔄 Sub complete\n"
        f"**Out:** {member_out.display_name}\n"
        f"**In:** {member_in.display_name}{voice_note}\n"
        f"**Team:** {vc.name if vc else out_vc_id}\n\n"
        f"{build_team_summary(guild, gid)}",
        view=DismissView(), ephemeral=False)


@bot.tree.command(name="import_mmr",
    description="[Admin] Import stats from the repo Excel file, or upload a new one.")
@is_admin()
async def import_mmr(interaction: discord.Interaction,
                     file: discord.Attachment = None):
    await interaction.response.defer(ephemeral=True)
    try:
        import openpyxl
        if file is not None:
            if not file.filename.endswith((".xlsx", ".csv")):
                await interaction.followup.send(
                    "⚠️ Please upload a `.xlsx` file.", ephemeral=True)
                return
            wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
            source_label = f"uploaded file `{file.filename}`"
        elif os.path.exists(STATS_FILE):
            wb = openpyxl.load_workbook(STATS_FILE)
            source_label = f"`{STATS_FILE}` from repo"
        else:
            await interaction.followup.send(
                f"⚠️ No file uploaded and `{STATS_FILE}` not found in the repo.\n"
                "Commit the Excel file to GitHub or upload it directly.",
                ephemeral=True)
            return

        gid = str(interaction.guild_id)
        if gid not in mmr_data:
            mmr_data[gid] = {}

        skip = ("collective", "summary")
        session_sheets = [s for s in wb.sheetnames
                          if not any(k in s.lower() for k in skip)
                          and s.lower() != "leaderboard"]

        per_player: dict = {}
        for sheet_name in session_sheets:
            players = parse_session_sheet(wb[sheet_name])
            if not players:
                continue
            players = calculate_mmr(players)
            players_sorted = sorted(players, key=lambda x: x["mmr"], reverse=True)
            session_ranks  = {p["name"]: i+1 for i, p in enumerate(players_sorted)}
            session_size   = len(players)
            for p in players:
                cname = p["name"]
                if cname not in per_player:
                    per_player[cname] = []
                per_player[cname].append({
                    "session":      sheet_name,
                    "mmr":          p["mmr"],
                    "kills":        p.get("kills", 0),
                    "assists":      p.get("assists", 0),
                    "deaths":       p.get("deaths", 0),
                    "kd":           p["kd"],
                    "captures":     p.get("captures", 0),
                    "obj_time":     p.get("obj_time", 0),
                    "points":       p["points"],
                    "session_rank": session_ranks.get(p["name"], "?"),
                    "session_size": session_size,
                })

        overall_mmr: dict = {}
        if "Leaderboard" in wb.sheetnames:
            lb_players = parse_leaderboard_sheet(wb["Leaderboard"])
            if lb_players:
                lb_players = calculate_mmr(lb_players)
                for p in lb_players:
                    overall_mmr[p["name"]] = p

        if not per_player and not overall_mmr:
            await interaction.followup.send("⚠️ No valid data found.", ephemeral=True)
            return

        all_names = set(per_player.keys()) | set(overall_mmr.keys())
        imported  = []
        for cname in all_names:
            existing      = mmr_data[gid].get(cname, {})
            sessions_list = per_player.get(cname, [])
            lb = overall_mmr.get(cname) or next(
                (v for k, v in overall_mmr.items() if k.lower() == cname.lower()), None)
            existing_sessions = [h["session"] for h in existing.get("history", [])]
            new_history = existing.get("history", [])
            for s in sessions_list:
                if s["session"] not in existing_sessions:
                    new_history.append(s)
            if lb:
                overall = lb["mmr"]
                kd, kills, deaths = lb["kd"], lb.get("kills", 0), lb.get("deaths", 0)
                points, obj_time  = lb["points"], lb["obj_time"]
                assists, captures = lb["assists"], lb["captures"]
                session_count     = lb["sessions"]
            elif sessions_list:
                overall = round(sum(s["mmr"] for s in sessions_list) / len(sessions_list), 1)
                last    = sessions_list[-1]
                kd, kills = last["kd"], last.get("kills", 0)
                deaths    = last.get("deaths", 0)
                points, obj_time, assists, captures = last["points"], 0, 0, 0
                session_count = len(sessions_list)
            else:
                continue
            gamertag = lb.get("gamertag", "") if lb else ""
            mmr_data[gid][cname] = {
                "mmr": overall, "kd": kd, "kills": kills, "deaths": deaths,
                "points": points, "obj_time": obj_time, "assists": assists,
                "captures": captures, "sessions": session_count,
                "gamertag": gamertag, "history": new_history,
            }
            imported.append((cname, overall, session_count))

        save_json(MMR_FILE, mmr_data)
        imported.sort(key=lambda x: x[1], reverse=True)

        lines = []
        for cname, mmr, sessions in imported:
            rname, ename = halo_rank(mmr)
            remoji = get_emoji(interaction.guild, ename)
            prov   = "*" if sessions < PROVISIONAL_SESSIONS else ""
            lines.append(f"{remoji} **{cname}**{prov} — {mmr} MMR")

        header = (
            f"✅ Imported **{len(imported)}** players from {source_label}!\n"
            f"_* = Provisional (fewer than {PROVISIONAL_SESSIONS} sessions)_\n"
        )
        await send_single_or_chunked(interaction, lines, header=header, ephemeral=True)

    except Exception as e:
        await interaction.followup.send(f"❌ Error: {e}", ephemeral=True)


@bot.tree.command(name="leaderboard",
    description="Show the Halo Reach MMR leaderboard.")
async def leaderboard(interaction: discord.Interaction):
    gmmr = get_guild_mmr(interaction.guild_id)
    if not gmmr:
        await interaction.response.send_message(
            "⚠️ No MMR data yet. An admin needs to run `/import_mmr` first.",
            ephemeral=True)
        return
    await interaction.response.defer()
    sorted_players = sorted(gmmr.items(), key=lambda x: x[1].get("mmr", 0), reverse=True)
    rank_groups: OrderedDict = OrderedDict()
    for pos, (name, data) in enumerate(sorted_players, 1):
        mmr      = data.get("mmr", 0)
        sessions = data.get("sessions", 0)
        rname, ename = halo_rank(mmr)
        prov     = "*" if sessions < PROVISIONAL_SESSIONS else ""
        gamertag = data.get("gamertag", "")
        gt_part  = f" {gamertag}" if gamertag else ""
        entry    = f"  `#{pos}` **{name}**{prov}{gt_part} — {mmr} MMR"
        if rname not in rank_groups:
            rank_groups[rname] = {"ename": ename, "entries": []}
        rank_groups[rname]["entries"].append(entry)
    lines = []
    for rname, group in rank_groups.items():
        remoji = get_emoji(interaction.guild, group["ename"])
        lines.append(f"{remoji} **{rname}**")
        lines.extend(group["entries"])
    header = (
        f"🏆 **Halo Night MMR Leaderboard**\n"
        f"_* = Provisional (fewer than {PROVISIONAL_SESSIONS} sessions)_\n"
    )
    # Leaderboard — no timeout (persistent public message)
    await send_single_or_chunked(interaction, lines, header=header, ephemeral=False)


@bot.tree.command(name="mmr",
    description="Look up a player's MMR, rank, and session history.")
async def mmr_lookup(interaction: discord.Interaction, player: str):
    gmmr  = get_guild_mmr(interaction.guild_id)
    match = next((v for k, v in gmmr.items() if k.lower() == player.lower()), None)
    name  = next((k for k in gmmr if k.lower() == player.lower()), player)
    if not match:
        await interaction.response.send_message(
            f"⚠️ No MMR data found for **{player}**.", ephemeral=True)
        return
    await interaction.response.defer()
    sorted_all    = sorted(gmmr.values(), key=lambda x: x.get("mmr", 0), reverse=True)
    total_players = len(gmmr)
    rank_pos      = next(
        (i+1 for i, p in enumerate(sorted_all) if p.get("mmr") == match.get("mmr")), "?")
    mmr           = match.get("mmr", 0)
    sessions      = match.get("sessions", 0)
    rname, ename  = halo_rank(mmr)
    remoji        = get_emoji(interaction.guild, ename)
    prov          = " *" if sessions < PROVISIONAL_SESSIONS else ""
    prov_note     = (f"\n_* Provisional — needs {PROVISIONAL_SESSIONS - sessions} "
                     f"more session(s) to confirm rank._"
                     if sessions < PROVISIONAL_SESSIONS else "")
    kills   = match.get("kills", 0)
    deaths  = match.get("deaths", 0)
    assists = match.get("assists", 0)
    kda     = round((kills + assists) / max(deaths, 1), 2)
    lines   = [
        f"**{name}** {remoji} *{rname}*{prov} — Rank **#{rank_pos} / {total_players}**{prov_note}",
        f"Overall MMR: **{mmr}** | Sessions: **{sessions}**",
        (f"Kills: {kills} | Deaths: {deaths} | K/D: {match.get('kd','?')} | KDA: {kda} | "
         f"Assists: {assists} | Points: {match.get('points','?')} | "
         f"Obj Time: {match.get('obj_time','?')}s | Captures: {match.get('captures','?')}"),
    ]
    history = match.get("history", [])
    if history:
        lines.append("\n📈 **Session Breakdown**")
        prev_mmr = None
        for h in history:
            h_rname, h_ename = halo_rank(h["mmr"])
            h_remoji  = get_emoji(interaction.guild, h_ename)
            s_rank    = h.get("session_rank", "?")
            s_size    = h.get("session_size", "?")
            arrow     = ("" if prev_mmr is None
                         else " ▲" if h["mmr"] > prev_mmr
                         else " ▼" if h["mmr"] < prev_mmr else " ─")
            lines.append(
                f"> {h_remoji} *{h_rname}* | **{h['session']}**: "
                f"{h['mmr']} MMR — #{s_rank}/{s_size}{arrow}")
            prev_mmr = h["mmr"]
    await send_single_or_chunked(interaction, lines, ephemeral=False, timeout=TIMEOUT_STAT)


@bot.tree.command(name="rank", description="Check your own current rank and MMR.")
async def rank(interaction: discord.Interaction):
    gmmr  = get_guild_mmr(interaction.guild_id)
    dname = interaction.user.display_name
    cname = canonical_name(dname)
    match = (gmmr.get(cname) or gmmr.get(dname)
             or next((v for k, v in gmmr.items() if k.lower() == cname.lower()), None))
    if not match:
        await interaction.response.send_message(
            f"⚠️ No MMR data found for **{dname}**.\n"
            "Your Discord display name needs to match your name in the spreadsheet.",
            ephemeral=True)
        return
    sorted_all    = sorted(gmmr.values(), key=lambda x: x.get("mmr", 0), reverse=True)
    total_players = len(gmmr)
    rank_pos      = next(
        (i+1 for i, p in enumerate(sorted_all) if p.get("mmr") == match.get("mmr")), "?")
    mmr           = match.get("mmr", 0)
    sessions      = match.get("sessions", 0)
    rname, ename  = halo_rank(mmr)
    remoji        = get_emoji(interaction.guild, ename)
    prov          = " *" if sessions < PROVISIONAL_SESSIONS else ""
    prov_note     = (f"\n_* {PROVISIONAL_SESSIONS - sessions} more session(s) "
                     f"until your rank is confirmed._"
                     if sessions < PROVISIONAL_SESSIONS else "")
    rk   = match.get("kills", 0)
    rd   = match.get("deaths", 0)
    ra   = match.get("assists", 0)
    rkda = round((rk + ra) / max(rd, 1), 2)
    msg  = (
        f"**{dname}** {remoji} *{rname}*{prov} — Rank **#{rank_pos} / {total_players}**{prov_note}\n"
        f"MMR: **{mmr}** | Sessions: **{sessions}**\n"
        f"Kills: {rk} | Deaths: {rd} | K/D: {match.get('kd','?')} | KDA: {rkda} | "
        f"Assists: {ra} | Points: {match.get('points','?')} | "
        f"Obj Time: {match.get('obj_time','?')}s | Captures: {match.get('captures','?')}"
    )
    await send_minimal(interaction, msg, ephemeral=True, timeout=TIMEOUT_STAT)


@bot.tree.command(name="compare", description="Compare 2 to 4 players side by side.")
async def compare(interaction: discord.Interaction,
                  player1: str, player2: str,
                  player3: str = None, player4: str = None):
    gmmr     = get_guild_mmr(interaction.guild_id)
    names_in = [player1, player2] + [p for p in [player3, player4] if p]
    players  = []
    for pname in names_in:
        d = next((v for k, v in gmmr.items() if k.lower() == pname.lower()), None)
        n = next((k for k in gmmr if k.lower() == pname.lower()), pname)
        if not d:
            await interaction.response.send_message(
                f"⚠️ No data found for **{pname}**.", ephemeral=True)
            return
        players.append((n, d))

    await interaction.response.defer()
    sorted_all = sorted(gmmr.values(), key=lambda x: x.get("mmr", 0), reverse=True)
    total      = len(gmmr)

    def rank_pos(d):
        return next(
            (i+1 for i, p in enumerate(sorted_all) if p.get("mmr") == d.get("mmr")), "?")

    def winner_arrow(vals, higher_is_better=True):
        try:
            floats = [float(v) for v in vals]
            best   = max(floats) if higher_is_better else min(floats)
            return [("🟰" if floats.count(best) > 1 else "⬆️") if f == best else ""
                    for f in floats]
        except:
            return [""] * len(vals)

    def kda(d):
        k = float(d.get("kills", 0))
        a = float(d.get("assists", 0))
        dth = max(float(d.get("deaths", 1)), 1)
        return round((k + a) / dth, 2)

    STATS = [
        ("Rank",     lambda d: f"#{rank_pos(d)}/{total}",            None),
        ("MMR",      lambda d: d.get("mmr", "?"),                     True),
        ("Kills",    lambda d: d.get("kills", "?"),                   True),
        ("Deaths",   lambda d: d.get("deaths", "?"),                  False),
        ("K/D",      lambda d: d.get("kd", "?"),                      True),
        ("KDA",      lambda d: kda(d),                                True),
        ("Assists",  lambda d: d.get("assists", "?"),                 True),
        ("Points",   lambda d: d.get("points", "?"),                  True),
        ("Obj Time", lambda d: f"{d.get('obj_time','?')}s",           True),
        ("Captures", lambda d: d.get("captures", "?"),                True),
        ("Sessions", lambda d: d.get("sessions", "?"),                None),
    ]

    col  = 11
    vcol = max(12, max(len(n) + 2 for n, _ in players))
    lines = []
    rank_headers = []
    for n, d in players:
        rname, ename = halo_rank(d.get("mmr", 0))
        remoji = get_emoji(interaction.guild, ename)
        rank_headers.append(f"**{n}** {remoji} *{rname}*")
    lines.append(" vs ".join(rank_headers) + "\n")
    lines.append(f"`{'Stat':<{col}}" + "".join(f"{'▸ '+n:<{vcol}}" for n, _ in players) + "`")
    lines.append(f"`{'-'*(col + vcol*len(players))}`")
    for label, fn, hib in STATS:
        vals   = [str(fn(d)) for _, d in players]
        arrows = winner_arrow(vals, hib) if hib is not None else [""] * len(players)
        row    = f"`{label:<{col}}" + "".join(f"{v:<{vcol}}" for v in vals) + "`"
        row   += " " + " ".join(a for a in arrows if a)
        lines.append(row)

    await send_single_or_chunked(interaction, lines, ephemeral=False, timeout=TIMEOUT_STAT)


@bot.tree.command(name="rivals",
    description="Head-to-head session history between two players.")
async def rivals(interaction: discord.Interaction, player1: str, player2: str):
    gmmr = get_guild_mmr(interaction.guild_id)
    d1   = next((v for k, v in gmmr.items() if k.lower() == player1.lower()), None)
    d2   = next((v for k, v in gmmr.items() if k.lower() == player2.lower()), None)
    n1   = next((k for k in gmmr if k.lower() == player1.lower()), player1)
    n2   = next((k for k in gmmr if k.lower() == player2.lower()), player2)
    if not d1:
        await interaction.response.send_message(
            f"⚠️ No data for **{player1}**.", ephemeral=True); return
    if not d2:
        await interaction.response.send_message(
            f"⚠️ No data for **{player2}**.", ephemeral=True); return
    await interaction.response.defer()
    h1     = {h["session"]: h for h in d1.get("history", [])}
    h2     = {h["session"]: h for h in d2.get("history", [])}
    shared = sorted(set(h1.keys()) & set(h2.keys()))
    if not shared:
        await interaction.followup.send(
            f"⚠️ **{n1}** and **{n2}** have no sessions in common.",
            ephemeral=True)
        return
    p1_wins, p2_wins, draws = 0, 0, 0
    lines = [f"⚔️ **{n1}** vs **{n2}** — {len(shared)} shared session(s)\n"]
    for session in shared:
        s1 = h1[session]; s2 = h2[session]
        _, e1 = halo_rank(s1["mmr"]); _, e2 = halo_rank(s2["mmr"])
        em1 = get_emoji(interaction.guild, e1)
        em2 = get_emoji(interaction.guild, e2)
        if s1["mmr"] > s2["mmr"]:
            winner = f"→ **{n1}** wins"; p1_wins += 1
        elif s2["mmr"] > s1["mmr"]:
            winner = f"→ **{n2}** wins"; p2_wins += 1
        else:
            winner = "→ Draw"; draws += 1
        r1 = f"#{s1.get('session_rank','?')}/{s1.get('session_size','?')}"
        r2 = f"#{s2.get('session_rank','?')}/{s2.get('session_size','?')}"
        lines.append(
            f"**{session}**\n"
            f"> {em1} {n1}: {s1['mmr']} MMR ({r1})\n"
            f"> {em2} {n2}: {s2['mmr']} MMR ({r2})\n"
            f"> {winner}")
    lines.append(
        f"\n🏆 **Head-to-head:** {n1} {p1_wins} — {p2_wins} {n2}"
        + (f" ({draws} draw)" if draws else ""))
    await send_single_or_chunked(interaction, lines, ephemeral=False, timeout=TIMEOUT_STAT)


@bot.tree.command(name="stats", description="Show top performers by stat category.")
async def stats(interaction: discord.Interaction):
    gmmr = get_guild_mmr(interaction.guild_id)
    if not gmmr:
        await interaction.response.send_message(
            "⚠️ No MMR data yet.", ephemeral=True); return
    await interaction.response.defer()
    categories = [
        ("mmr",      "🏆 Top MMR",        True),
        ("kd",       "🎯 Best K/D",        True),
        ("kills",    "💀 Most Kills",      True),
        ("deaths",   "☠️ Fewest Deaths",   False),
        ("assists",  "🤝 Most Assists",    True),
        ("obj_time", "⏱️ Most Obj Time",   True),
        ("captures", "🚩 Most Captures",   True),
        ("points",   "⭐ Most Points",     True),
    ]
    lines = ["📊 **Stat Leaders**\n"]
    for key, label, hib in categories:
        valid = [(n, d) for n, d in gmmr.items() if d.get(key) is not None]
        if not valid:
            continue
        best_name, best_data = sorted(valid, key=lambda x: x[1].get(key, 0), reverse=hib)[0]
        val    = best_data.get(key, "?")
        _, ename = halo_rank(best_data.get("mmr", 0))
        remoji = get_emoji(interaction.guild, ename)
        suffix = "s" if key == "obj_time" else ""
        lines.append(f"{label}: **{best_name}** {remoji} — {val}{suffix}")
    await send_single_or_chunked(interaction, lines, ephemeral=False, timeout=TIMEOUT_STAT)


@bot.tree.command(name="session",
    description="Look up a player's stats from a specific session number.")
async def session_lookup(interaction: discord.Interaction, number: int, player: str):
    gmmr  = get_guild_mmr(interaction.guild_id)
    match = next((v for k, v in gmmr.items() if k.lower() == player.lower()), None)
    name  = next((k for k in gmmr if k.lower() == player.lower()), player)
    if not match:
        await interaction.response.send_message(
            f"⚠️ No data found for **{player}**.", ephemeral=True); return
    history     = match.get("history", [])
    session_key = f"Session {number}"
    entry       = next(
        (h for h in history if h["session"].lower() == session_key.lower()), None)
    if not entry:
        available = ", ".join(f"`{h['session']}`" for h in history) or "none"
        await interaction.response.send_message(
            f"⚠️ **{name}** has no data for Session {number}.\nAvailable: {available}",
            ephemeral=True)
        return
    mmr    = entry["mmr"]
    s_rank = entry.get("session_rank", "?")
    s_size = entry.get("session_size", "?")
    rname, ename = halo_rank(mmr)
    remoji = get_emoji(interaction.guild, ename)
    kills  = entry.get("kills", "?")
    deaths = entry.get("deaths", "?")
    kd     = entry.get("kd", "?")
    points = entry.get("points", "?")
    asst   = entry.get("assists", "?")
    caps   = entry.get("captures", "?")
    try:
        kda = round((float(kills) + float(asst)) / max(float(deaths), 1), 2)
    except:
        kda = "?"
    lines = [
        f"**{name}** — {entry['session']}",
        f"{remoji} *{rname}* | MMR: **{mmr}** | Session Rank: **#{s_rank}/{s_size}**",
        f"Kills: {kills} | Deaths: {deaths} | K/D: {kd} | KDA: {kda}",
        f"Assists: {asst} | Points: {points} | Captures: {caps}",
    ]
    await send_minimal(interaction, "\n".join(lines), ephemeral=False, timeout=TIMEOUT_STAT)


@bot.tree.command(name="export", description="[Admin] Download the MMR data file.")
@is_admin()
async def export(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    gid  = str(interaction.guild_id)
    gmmr = mmr_data.get(gid, {})
    if not gmmr:
        await interaction.followup.send("⚠️ No MMR data to export.", ephemeral=True)
        return
    data   = json.dumps({gid: gmmr}, indent=2)
    buffer = io.BytesIO(data.encode())
    fname  = f"mmr_data_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M')}.json"
    await interaction.followup.send(
        "📦 Here is your current MMR data file:",
        file=discord.File(buffer, filename=fname),
        ephemeral=True)


@bot.tree.command(name="presets", description="[Admin] View and load saved team presets.")
@is_admin()
async def view_presets(interaction: discord.Interaction):
    if not presets.get(str(interaction.guild_id)):
        await interaction.response.send_message(
            "⚠️ No presets saved yet. Use 💾 Save Preset in `/teams`.", ephemeral=True)
        return
    gp = presets.get(str(interaction.guild_id), {})
    options = [discord.SelectOption(label=n,
               description=d.get("note","")[:50] or "No notes", value=n)
               for n, d in list(gp.items())[:25]]
    view = discord.ui.View(timeout=TIMEOUT_MENU)
    select = discord.ui.Select(placeholder="Choose a preset to view...",
                               options=options, row=0)
    async def on_select(inter: discord.Interaction):
        preset = gp.get(select.values[0])
        if not preset:
            await inter.response.edit_message(content="⚠️ Not found.", view=view)
            return
        lines = [f"📋 **{select.values[0]}**"]
        if preset.get("note"): lines.append(f"_{preset['note']}_")
        for vc_name, members in preset["teams"].items():
            lines.append(f"**{vc_name}**: {', '.join(members)}")
        await inter.response.edit_message(content="\n".join(lines), view=view)
    select.callback = on_select
    view.add_item(select)
    await interaction.response.send_message(
        "📋 **Saved Presets** — select one to view:",
        view=view, ephemeral=True)


@bot.tree.command(name="history", description="[Admin] View past team configurations.")
@is_admin()
async def view_history(interaction: discord.Interaction):
    if not team_history.get(str(interaction.guild_id)):
        await interaction.response.send_message(
            "⚠️ No team history yet.", ephemeral=True)
        return
    await interaction.response.send_message(
        "📜 **Team History** — select an entry:",
        view=TeamHistoryView(interaction.guild), ephemeral=True)


@bot.tree.command(name="podium", description="Show 🥇🥈🥉 top 3 in every stat category.")
async def podium(interaction: discord.Interaction):
    gmmr = get_guild_mmr(interaction.guild_id)
    if not gmmr:
        await interaction.response.send_message(
            "⚠️ No MMR data yet.", ephemeral=True)
        return
    await interaction.response.defer()

    categories = [
        ("mmr",      "🏆 MMR",         True),
        ("kd",       "🎯 K/D",          True),
        ("kills",    "💀 Kills",        True),
        ("deaths",   "☠️ Fewest Deaths", False),
        ("assists",  "🤝 Assists",      True),
        ("obj_time", "⏱️ Obj Time",     True),
        ("captures", "🚩 Captures",     True),
        ("points",   "⭐ Points",       True),
    ]

    medals = ["🥇", "🥈", "🥉"]
    lines  = ["🏅 **Halo Night Podium**\n"]

    for key, label, higher_is_better in categories:
        valid = [(n, d) for n, d in gmmr.items() if d.get(key) is not None]
        if not valid:
            continue
        sorted_valid = sorted(valid, key=lambda x: x[1].get(key, 0),
                              reverse=higher_is_better)[:3]
        suffix = "s" if key == "obj_time" else ""
        lines.append(f"**{label}**")
        for i, (name, data) in enumerate(sorted_valid):
            val    = data.get(key, "?")
            _, ename = halo_rank(data.get("mmr", 0))
            remoji = get_emoji(interaction.guild, ename)
            lines.append(f"{medals[i]} {remoji} **{name}** — {val}{suffix}")
        lines.append("")

    await send_single_or_chunked(interaction, lines, ephemeral=False, timeout=TIMEOUT_STAT)


@bot.tree.command(
    name="mmr_hub",
    description="Open a one-command MMR/stats interface with a dropdown menu.")
async def mmr_hub(interaction: discord.Interaction):
    handlers = {
        "leaderboard": lambda inter: leaderboard.callback(inter),
        "mmr": lambda inter, player: mmr_lookup.callback(inter, player),
        "compare": lambda inter, players: compare.callback(
            inter,
            players[0],
            players[1],
            players[2] if len(players) > 2 else None,
            players[3] if len(players) > 3 else None,
        ),
        "rivals": lambda inter, p1, p2: rivals.callback(inter, p1, p2),
        "session": lambda inter, number, player: session_lookup.callback(inter, number, player),
        "podium": lambda inter: podium.callback(inter),
        "stats": lambda inter: stats.callback(inter),
    }
    await interaction.response.send_message(
        "📘 **MMR & Stats Interface** — select an action from the dropdown:",
        view=MMRHubView(handlers),
        ephemeral=True,
    )


@bot.tree.command(name="help", description="List all available commands.")
async def help_command(interaction: discord.Interaction):
    is_admin_user = interaction.user.guild_permissions.administrator
    # get_commands() alone only returns global commands registered before sync.
    # Fetch from the guild tree which is always populated after on_ready sync.
    guild_obj = interaction.guild
    all_cmds  = sorted(
        bot.tree.get_commands(guild=guild_obj) or bot.tree.get_commands(),
        key=lambda c: c.name
    )
    admin_cmds, everyone_cmds = [], []
    for cmd in all_cmds:
        if cmd.name == "help":
            continue
        desc = cmd.description or "No description."
        if desc.startswith("[Admin]"):
            admin_cmds.append((f"`/{cmd.name}`", desc.replace("[Admin] ", "")))
        else:
            everyone_cmds.append((f"`/{cmd.name}`", desc))
    lines = ["📖 **Halo Night Bot — Command List**\n", "**Everyone**"]
    for cmd, desc in everyone_cmds:
        lines.append(f"> {cmd} — {desc}")
    if is_admin_user:
        lines.append("\n**Admin Only**")
        for cmd, desc in admin_cmds:
            lines.append(f"> {cmd} — {desc}")
    else:
        lines.append("\n_Admin commands are hidden. Ask a server admin for help._")
    await send_minimal(interaction, "\n".join(lines), ephemeral=True, timeout=TIMEOUT_MENU)


@bot.tree.command(name="sync",
    description="[Admin] Force sync slash commands if any are missing.")
@is_admin()
async def sync_commands(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    await bot.tree.sync()
    for guild in bot.guilds:
        try:
            await bot.tree.sync(guild=guild)
        except Exception as e:
            log.warning(
                "Guild command sync failed for %s (%s): %s",
                getattr(guild, "name", "?"),
                guild.id,
                e,
            )
    await interaction.followup.send(
        "✅ Commands synced! New commands should appear within 30 seconds.",
        ephemeral=True)


async def _admin_error(interaction: discord.Interaction, error):
    if isinstance(error, app_commands.CheckFailure):
        await interaction.response.send_message(
            "❌ Administrator permissions required.", ephemeral=True)
    else:
        raise error


# ─────────────────────────────────────────────
# ERROR HANDLERS
# ─────────────────────────────────────────────
@bot.tree.error
async def on_app_command_error(
    interaction: discord.Interaction, error: app_commands.AppCommandError
):
    if isinstance(error, app_commands.CommandInvokeError):
        log.exception("App command failed", exc_info=error.original)
        try:
            msg = "Something went wrong. If this keeps happening, notify a server admin."
            if interaction.response.is_done():
                await interaction.followup.send(msg, ephemeral=True)
            else:
                await interaction.response.send_message(msg, ephemeral=True)
        except discord.HTTPException:
            log.warning("Could not send error message for failed app command")
        return
    raise error


# ─────────────────────────────────────────────
# STARTUP
# ─────────────────────────────────────────────
@bot.event
async def on_ready():
    await bot.tree.sync()
    for guild in bot.guilds:
        try:
            await bot.tree.sync(guild=guild)
        except Exception as e:
            log.warning(
                "Guild command sync failed for %s (%s): %s",
                getattr(guild, "name", "?"),
                guild.id,
                e,
            )
    log.info("Logged in as %s (ID: %s)", bot.user, bot.user.id)
    log.info("Slash commands synced to %s guild(s).", len(bot.guilds))

if __name__ == "__main__":
    bot.run(TOKEN)
